from openpyxl import Workbook, load_workbook
import os
import tkinter as tk
import pandas as pd
import joblib

from sklearn.model_selection import train_test_split
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score


# ============================================================
# MACHINE LEARNING MODEL TRAINING
# ============================================================

def train_model():

    filename = "student data.xlsx"

    # Check whether Excel file exists
    if not os.path.exists(filename):

        print("================================")
        print("Excel file not found.")
        print("Please create student data.xlsx")
        print("and add training data first.")
        print("================================")

        return None

    # Read Excel file
    data = pd.read_excel(filename)

    print("\n================================")
    print("       DATASET INFORMATION")
    print("================================")

    print("Total rows:", len(data))
    print(data.head())

    # Required columns
    required_columns = [
        "attendance",
        "study_hours",
        "internal_marks",
        "assignment",
        "previous_score",
        "performance"
    ]

    # Check columns
    for column in required_columns:

        if column not in data.columns:

            print("Missing column:", column)

            return None

    # Remove empty rows
    data = data.dropna(
        subset=required_columns
    )

    # --------------------------------------------------------
    # Check performance values
    # --------------------------------------------------------

    valid_performance = [
        "Good",
        "Average",
        "Poor"
    ]

    data = data[
        data["performance"].isin(
            valid_performance
        )
    ]

    # Need at least 3 classes
    if data["performance"].nunique() < 3:

        print("\nError:")
        print("Performance column must contain")
        print("Good, Average and Poor.")

        return None

    # --------------------------------------------------------
    # X = INPUT FEATURES
    # --------------------------------------------------------

    X = data[
        [
            "attendance",
            "study_hours",
            "internal_marks",
            "assignment",
            "previous_score"
        ]
    ]

    # --------------------------------------------------------
    # y = TARGET
    # --------------------------------------------------------

    y = data["performance"]

    print("\nInput features:")
    print(X.columns.tolist())

    print("\nTarget:")
    print("performance")

    # --------------------------------------------------------
    # 80% TRAINING / 20% TESTING
    # --------------------------------------------------------

    X_train, X_test, y_train, y_test = train_test_split(

        X,
        y,

        test_size=0.20,

        random_state=42,

        stratify=y
    )

    print("\n================================")
    print("          DATA SPLIT")
    print("================================")

    print(
        "Training rows:",
        len(X_train)
    )

    print(
        "Testing rows:",
        len(X_test)
    )

    # ========================================================
    # CREATE MODELS
    # ========================================================

    decision_tree = DecisionTreeClassifier(
        random_state=42
    )

    random_forest = RandomForestClassifier(
        n_estimators=100,
        random_state=42
    )

    logistic_regression = LogisticRegression(
        max_iter=1000
    )

    # ========================================================
    # TRAIN MODELS
    # ========================================================

    print("\n================================")
    print("       TRAINING MODELS")
    print("================================")

    decision_tree.fit(
        X_train,
        y_train
    )

    print(
        "Decision Tree trained"
    )

    random_forest.fit(
        X_train,
        y_train
    )

    print(
        "Random Forest trained"
    )

    logistic_regression.fit(
        X_train,
        y_train
    )

    print(
        "Logistic Regression trained"
    )

    # ========================================================
    # PREDICTIONS
    # ========================================================

    dt_prediction = decision_tree.predict(
        X_test
    )

    rf_prediction = random_forest.predict(
        X_test
    )

    lr_prediction = logistic_regression.predict(
        X_test
    )

    # ========================================================
    # ACCURACY
    # ========================================================

    dt_accuracy = accuracy_score(
        y_test,
        dt_prediction
    )

    rf_accuracy = accuracy_score(
        y_test,
        rf_prediction
    )

    lr_accuracy = accuracy_score(
        y_test,
        lr_prediction
    )

    print("\n================================")
    print("        MODEL ACCURACY")
    print("================================")

    print(
        "Decision Tree:",
        round(dt_accuracy * 100, 2),
        "%"
    )

    print(
        "Random Forest:",
        round(rf_accuracy * 100, 2),
        "%"
    )

    print(
        "Logistic Regression:",
        round(lr_accuracy * 100, 2),
        "%"
    )

    # ========================================================
    # FIND BEST MODEL
    # ========================================================

    models = {

        "Decision Tree": (
            decision_tree,
            dt_accuracy
        ),

        "Random Forest": (
            random_forest,
            rf_accuracy
        ),

        "Logistic Regression": (
            logistic_regression,
            lr_accuracy
        )
    }

    best_name, (
        best_model,
        best_accuracy
    ) = max(
        models.items(),
        key=lambda item: item[1][1]
    )

    print("\n================================")
    print("          BEST MODEL")
    print("================================")

    print(
        "Best model:",
        best_name
    )

    print(
        "Best accuracy:",
        round(best_accuracy * 100, 2),
        "%"
    )

    # ========================================================
    # SAVE BEST MODEL
    # ========================================================

    model_filename = (
        "student_performance_model.pkl"
    )

    joblib.dump(
        best_model,
        model_filename
    )

    print(
        "Model saved:",
        model_filename
    )

    print("================================")

    return best_model


# ============================================================
# TRAIN MODEL BEFORE GUI
# ============================================================

model = train_model()


# ============================================================
# TKINTER WINDOW
# ============================================================

root = tk.Tk()

root.title(
    "Smart Student Performance Prediction System"
)

root.geometry(
    "1000x650"
)

root.configure(
    bg="#EAF2F8"
)


# ============================================================
# CLEAR FUNCTION
# ============================================================

def clear():

    stuid.delete(
        0,
        tk.END
    )

    studentname.delete(
        0,
        tk.END
    )

    attid.delete(
        0,
        tk.END
    )

    sthrid.delete(
        0,
        tk.END
    )

    iatid.delete(
        0,
        tk.END
    )

    assid.delete(
        0,
        tk.END
    )

    previd.delete(
        0,
        tk.END
    )

    prediction.config(
        text="____________________________"
    )

    risk.config(
        text="____________________________"
    )

    result.config(
        text="____________________________"
    )


# ============================================================
# PREDICT FUNCTION
# ============================================================

def predict():

    try:

        attendance = float(
            attid.get()
        )

        studyhours = float(
            sthrid.get()
        )

        internal_marks = float(
            iatid.get()
        )

        assignment = float(
            assid.get()
        )

        previous_score = float(
            previd.get()
        )

        # ----------------------------------------------------
        # VALIDATION
        # ----------------------------------------------------

        if not (
            0 <= attendance <= 100
        ):

            prediction.config(
                text="Invalid attendance"
            )

            return

        if not (
            0 <= studyhours <= 8
        ):

            prediction.config(
                text="Study hours must be 0-8"
            )

            return

        if not (
            0 <= internal_marks <= 100
        ):

            prediction.config(
                text="Invalid IAT marks"
            )

            return

        if not (
            0 <= assignment <= 100
        ):

            prediction.config(
                text="Invalid assignment"
            )

            return

        if not (
            0 <= previous_score <= 100
        ):

            prediction.config(
                text="Invalid previous score"
            )

            return

        # ----------------------------------------------------
        # CHECK MODEL
        # ----------------------------------------------------

        if model is None:

            prediction.config(
                text="MODEL NOT AVAILABLE"
            )

            risk.config(
                text="Train model first"
            )

            result.config(
                text="Check student data.xlsx"
            )

            return

        # ----------------------------------------------------
        # CREATE NEW STUDENT DATA
        # ----------------------------------------------------

        new_student = pd.DataFrame(

            [
                [
                    attendance,
                    studyhours,
                    internal_marks,
                    assignment,
                    previous_score
                ]
            ],

            columns=[
                "attendance",
                "study_hours",
                "internal_marks",
                "assignment",
                "previous_score"
            ]
        )

        # ----------------------------------------------------
        # MACHINE LEARNING PREDICTION
        # ----------------------------------------------------

        prediction_value = model.predict(
            new_student
        )[0]

        # ----------------------------------------------------
        # DISPLAY RESULT
        # ----------------------------------------------------

        if prediction_value == "Good":

            prediction.config(
                text="GOOD PERFORMANCE"
            )

            risk.config(
                text="LOW RISK"
            )

            result.config(
                text="EXCELLENT! MAINTAIN PERFORMANCE"
            )

        elif prediction_value == "Average":

            prediction.config(
                text="AVERAGE PERFORMANCE"
            )

            risk.config(
                text="MEDIUM RISK"
            )

            result.config(
                text="YOU CAN IMPROVE YOUR PERFORMANCE"
            )

        else:

            prediction.config(
                text="POOR PERFORMANCE"
            )

            risk.config(
                text="HIGH RISK"
            )

            result.config(
                text="HIGH FOCUS ON PERFORMANCE"
            )

        # ----------------------------------------------------
        # TERMINAL OUTPUT
        # ----------------------------------------------------

        print(
            "\n------- STUDENT PERFORMANCE -------"
        )

        print(
            "Student Name =",
            studentname.get()
        )

        print(
            "Student ID =",
            stuid.get()
        )

        print(
            "Attendance =",
            attendance,
            "%"
        )

        print(
            "Study Hours =",
            studyhours
        )

        print(
            "Internal Marks =",
            internal_marks,
            "%"
        )

        print(
            "Assignment =",
            assignment,
            "%"
        )

        print(
            "Previous Score =",
            previous_score,
            "%"
        )

        print(
            "Prediction =",
            prediction_value
        )

        print(
            "----------------------------------"
        )

    except ValueError:

        prediction.config(
            text="INVALID INPUT"
        )

        risk.config(
            text="PLEASE ENTER NUMBERS"
        )

        result.config(
            text="CHECK ALL INPUT FIELDS"
        )


# ============================================================
# SAVE TO EXCEL
# ============================================================

def savetoexcel():

    studentid = stuid.get()

    student_name = studentname.get()

    attendance = attid.get()

    studyhours = sthrid.get()

    internalmarks = iatid.get()

    assignment = assid.get()

    previousscore = previd.get()

    filename = "student data.xlsx"

    # --------------------------------------------------------
    # CREATE NEW EXCEL FILE
    # --------------------------------------------------------

    if not os.path.exists(filename):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "student data"

        sheet.append(
            [
                "student_id",
                "student_name",
                "attendance",
                "study_hours",
                "internal_marks",
                "assignment",
                "previous_score",
                "performance"
            ]
        )

    # --------------------------------------------------------
    # OPEN EXISTING EXCEL FILE
    # --------------------------------------------------------

    else:

        workbook = load_workbook(
            filename
        )

        sheet = workbook[
            "student data"
        ]

    # --------------------------------------------------------
    # GET CURRENT PREDICTION
    # --------------------------------------------------------

    current_prediction = prediction.cget(
        "text"
    )

    if current_prediction == "GOOD PERFORMANCE":

        performance = "Good"

    elif current_prediction == "AVERAGE PERFORMANCE":

        performance = "Average"

    elif current_prediction == "POOR PERFORMANCE":

        performance = "Poor"

    else:

        performance = ""

    # --------------------------------------------------------
    # ADD DATA
    # --------------------------------------------------------

    sheet.append(
        [
            studentid,
            student_name,
            attendance,
            studyhours,
            internalmarks,
            assignment,
            previousscore,
            performance
        ]
    )

    workbook.save(
        filename
    )

    print(
        "Student data successfully saved in Excel"
    )

    print(
        "Excel location:",
        os.path.abspath(filename)
    )


# ============================================================
# TITLE
# ============================================================

label = tk.Label(

    root,

    text=(
        "SMART STUDENT PERFORMANCE\n"
        "PREDICTION SYSTEM"
    ),

    font=(
        "Arial",
        22,
        "bold"
    ),

    bg="#1F4E78",

    fg="white",

    pady=15
)

label.pack(
    fill="x",
    padx=20,
    pady=15
)


# ============================================================
# STUDENT INFORMATION
# ============================================================

frame1 = tk.LabelFrame(

    root,

    text="Student Information",

    font=(
        "Arial",
        12,
        "bold"
    ),

    padx=10,

    pady=10,

    bg="white",

    fg="black"
)

frame1.place(
    x=30,
    y=100,
    width=500,
    height=250
)


tk.Label(

    frame1,

    text="Student ID",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white",

    fg="black"

).grid(

    row=0,

    column=0,

    padx=15,

    pady=15,

    sticky="w"
)


stuid = tk.Entry(

    frame1,

    width=30,

    font=(
        "Arial",
        12,
        "bold"
    )
)

stuid.grid(
    row=0,
    column=1,
    padx=15,
    pady=15
)


tk.Label(

    frame1,

    text="Student Name",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white",

    fg="black"

).grid(

    row=1,

    column=0,

    padx=15,

    pady=15
)


studentname = tk.Entry(

    frame1,

    width=30,

    font=(
        "Arial",
        12,
        "bold"
    )
)

studentname.grid(
    row=1,
    column=1,
    padx=15,
    pady=15
)


# ============================================================
# ACADEMIC INFORMATION
# ============================================================

frame2 = tk.LabelFrame(

    root,

    text="Academic Information",

    font=(
        "Arial",
        12,
        "bold"
    ),

    padx=10,

    pady=10,

    bg="white",

    fg="black"
)

frame2.place(
    x=520,
    y=100,
    width=450,
    height=250
)


tk.Label(

    frame2,

    text="Attendance %",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white",

    fg="black"

).grid(

    row=0,

    column=0,

    padx=10,

    pady=8
)


attid = tk.Entry(

    frame2,

    width=25,

    font=(
        "Arial",
        12,
        "bold"
    )
)

attid.grid(
    row=0,
    column=1,
    padx=10,
    pady=8
)


tk.Label(

    frame2,

    text="Study Hours",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white",

    fg="black"

).grid(

    row=1,

    column=0,

    padx=10,

    pady=8
)


sthrid = tk.Entry(

    frame2,

    width=25,

    font=(
        "Arial",
        12,
        "bold"
    )
)

sthrid.grid(
    row=1,
    column=1,
    padx=10,
    pady=8
)


tk.Label(

    frame2,

    text="Internal Marks",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white",

    fg="black"

).grid(

    row=2,

    column=0,

    padx=10,

    pady=8
)


iatid = tk.Entry(

    frame2,

    width=25,

    font=(
        "Arial",
        12,
        "bold"
    )
)

iatid.grid(
    row=2,
    column=1,
    padx=10,
    pady=8
)


tk.Label(

    frame2,

    text="Assignment",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white",

    fg="black"

).grid(

    row=3,

    column=0,

    padx=10,

    pady=8
)


assid = tk.Entry(

    frame2,

    width=25,

    font=(
        "Arial",
        12,
        "bold"
    )
)

assid.grid(
    row=3,
    column=1,
    padx=10,
    pady=8
)


tk.Label(

    frame2,

    text="Previous Score",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white",

    fg="black"

).grid(

    row=4,

    column=0,

    padx=10,

    pady=8
)


previd = tk.Entry(

    frame2,

    width=25,

    font=(
        "Arial",
        12,
        "bold"
    )
)

previd.grid(
    row=4,
    column=1,
    padx=10,
    pady=8
)


# ============================================================
# BUTTON FRAME
# ============================================================

frame3 = tk.Frame(
    root,
    bg="lightgrey"
)

frame3.place(
    x=80,
    y=370,
    width=900,
    height=60
)


predictbutton = tk.Button(

    frame3,

    text="Predict Performance",

    width=18,

    height=2,

    font=(
        "Arial",
        11,
        "bold"
    ),

    bg="lightblue",

    command=predict
)

predictbutton.grid(
    row=0,
    column=0,
    padx=10,
    pady=6
)


clearbutton = tk.Button(

    frame3,

    text="Clear",

    width=12,

    height=2,

    font=(
        "Arial",
        11,
        "bold"
    ),

    bg="lightyellow",

    command=clear
)

clearbutton.grid(
    row=0,
    column=1,
    padx=10,
    pady=6
)


exitbutton = tk.Button(

    frame3,

    text="Exit",

    width=12,

    height=2,

    font=(
        "Arial",
        11,
        "bold"
    ),

    bg="red",

    fg="white",

    command=root.destroy
)

exitbutton.grid(
    row=0,
    column=2,
    padx=10,
    pady=6
)


savebutton = tk.Button(

    frame3,

    text="Save",

    width=12,

    height=2,

    font=(
        "Arial",
        11,
        "bold"
    ),

    bg="lightgreen",

    command=savetoexcel
)

savebutton.grid(
    row=0,
    column=3,
    padx=10,
    pady=6
)


# ============================================================
# PREDICTION RESULTS
# ============================================================

frame4 = tk.LabelFrame(

    root,

    text="Prediction Results",

    font=(
        "Arial",
        12,
        "bold"
    ),

    padx=20,

    pady=10,

    bg="white"
)

frame4.place(
    x=30,
    y=450,
    width=940,
    height=170
)


tk.Label(

    frame4,

    text="Prediction",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white"

).grid(

    row=0,

    column=0,

    padx=10,

    pady=10
)


prediction = tk.Label(

    frame4,

    text="____________________________",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white"
)

prediction.grid(
    row=0,
    column=1,
    padx=10,
    pady=10
)


tk.Label(

    frame4,

    text="Risk",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white"

).grid(

    row=1,

    column=0,

    padx=10,

    pady=10
)


risk = tk.Label(

    frame4,

    text="____________________________",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white"
)

risk.grid(
    row=1,
    column=1,
    padx=10,
    pady=10
)


tk.Label(

    frame4,

    text="Recommendations",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white"

).grid(

    row=2,

    column=0,

    padx=10,

    pady=10
)


result = tk.Label(

    frame4,

    text="____________________________",

    font=(
        "Arial",
        12,
        "bold"
    ),

    bg="white"
)

result.grid(
    row=2,
    column=1,
    padx=10,
    pady=10
)


# ============================================================
# START TKINTER
# ============================================================

root.mainloop()